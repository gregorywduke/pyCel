# pyCel GUI

import PySimpleGUI as sg
import openpyxl

from newXL import *

# Get data file
filename = sg.popup_get_file('Select Excel File to Process')
#output = sg.popup_get_folder('Select Folder to Output New Files')

# Load data into openpyxl format
xlFile = openpyxl.load_workbook(filename)
xlSheet = xlFile.active

# Create new empty file
wrFile = openpyxl.Workbook()
wrSheet = wrFile.active

# Get column count and first client
columns = xlSheet.max_row
client = xlSheet.cell(row = 2, column = 14)
clNext = xlSheet.cell(row = 3, column = 14)

# Transcribe labels from xlSheet to wrSheet
for x in range(1, columns):
    data1 = xlSheet.cell(row = 1, column = x)

    if data1 == 0:
        break
    else:
        wrSheet.cell(row = 1, column = x).value = data1.value


# Transcribe first client data
for x in range(1, columns):
    data1 = xlSheet.cell(row = 2, column = x)

    if data1 == 0:
        break
    else:
        wrSheet.cell(row = 2, column = x).value = data1.value



wrFile.save("client1.xlsx")




