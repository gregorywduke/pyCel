# newXL.py

import openpyxl
import datetime

from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
from openpyxl.drawing.image import Image

def getRows( filename ):

    # Load data into openpyxl format
    xlFile = openpyxl.load_workbook(filename)
    xlSheet = xlFile.active

    rowsNum = xlSheet.max_row
    return rowsNum

def getClientCol( filename ):
    # Load data into openpyxl format
    xlFile = openpyxl.load_workbook(filename)
    xlSheet = xlFile.active

    for x in range(1, 20):
        posClient = xlSheet.cell(row=1, column=x)
        if str(posClient.value) == 'Client':
            clientColumn = x
            return clientColumn

    print('Could not find client column')

# Creates one RECAP file per customer
def recap( filename, template, rowX, clNum, weekNum, dateRange, logo ):

    # Load data into openpyxl format
    xlFile = openpyxl.load_workbook(filename)
    xlSheet = xlFile.active

    # Open Template to Append
    wrFile = openpyxl.load_workbook(template)
    wrSheet = wrFile.active

    # Get column count and first client
    #columns = xlSheet.max_column
    columns = getClientCol( filename )
    client = xlSheet.cell(row=rowX, column=columns)
    rowY = rowX + 1
    clNext = xlSheet.cell(row=rowY, column=columns)
    rowY = 3
    rowCount = 2

    # Manage item totals
    bottlesSold = 0
    custEngage = 0

    # Loop processes each file
    while client != 0:
        # Transcribe from dataset to client file
        for x in range(1, columns):
            data1 = xlSheet.cell(row=rowX, column=x)

            # Calculates respective totals
            if x == 6:
                print('Data 1: ', data1.value)
                convert = int(data1.value)
                bottlesSold += convert
            if x == 8:
                convert = int(data1.value)
                custEngage += convert

            # If cell to be copied is empty
            if data1 == 0:
                break
            # Copy data to new sheet
            else:
                wrSheet.cell(row=rowY, column=x).value = data1.value
        rowY += 1
        rowCount += 1

        # If next client is empty, sheet is finished
        if clNext == 0:

            # Sort data values in client file by date
            sort = False
            # maxRow = wrSheet.max_row + 1
            # maxCol = wrSheet.max_column + 1
            row1 = 3
            row2 = 4
            col1 = 1
            swapCount = 0
            wrSheet.cell(row=row1, column=1)
            date1 = wrSheet.cell(row=row1, column=1).value
            date2 = wrSheet.cell(row=row2, column=1).value
            check = wrSheet.cell(row=row2, column=1)

            # SKIP SORT IF CLIENT IS ONLY 1 ROW
            if rowCount == 3:
                sort = True

            # MANUAL SORT IF CLIENT IS 2 ROWS
            elif rowCount == 4:
                sort = True
                da1 = date1.day
                da2 = date2.day

                # ROW SWAP
                if (da1 > da2):
                    swapCount += 1

                    # SWAP COLUMN DATETIMES
                    col1 = 1
                    temp = wrSheet.cell(row=row1, column=col1).value
                    wrSheet.cell(row=row1, column=col1).value = wrSheet.cell(row=row2, column=col1).value
                    wrSheet.cell(row=row2, column=col1).value = temp

                    # SWAP COLUMN STRINGS
                    for y in range(2, 10):
                        temp = wrSheet.cell(row=row1, column=y).value
                        wrSheet.cell(row=row1, column=y).value = wrSheet.cell(row=row2, column=y).value
                        wrSheet.cell(row=row2, column=y).value = temp

            # Bubble Sort
            while not sort:
                # PASS X
                for x in range(3, rowCount + 1):

                    # ITERATE & LOOK FOR ROWS TO SWAPS
                    if check.value is not None:
                        da1 = date1.day
                        da2 = date2.day

                        # ROW SWAP
                        if (da1 > da2):
                            swapCount += 1

                            # SWAP COLUMN DATETIMES
                            col1 = 1
                            temp = wrSheet.cell(row=row1, column=col1).value
                            wrSheet.cell(row=row1, column=col1).value = wrSheet.cell(row=row2, column=col1).value
                            wrSheet.cell(row=row2, column=col1).value = temp

                            # SWAP COLUMN STRINGS
                            for y in range(2, 10):
                                temp = wrSheet.cell(row=row1, column=y).value
                                wrSheet.cell(row=row1, column=y).value = wrSheet.cell(row=row2, column=y).value
                                wrSheet.cell(row=row2, column=y).value = temp

                        if (x == rowCount and swapCount == 0):
                            # If no swaps occured, sort is complete
                            sort = True

                        swapCount = 0
                        row1 += 1
                        row2 += 1
                        date1 = wrSheet.cell(row=row1, column=1).value
                        date2 = wrSheet.cell(row=row2, column=1).value
                        check = wrSheet.cell(row=row2, column=1)

                    # ELSE, NEW PASS (RESET SWAP SEARCH)
                    else:
                        row1 = 3
                        row2 = 4
                        date1 = wrSheet.cell(row=row1, column=1).value
                        date2 = wrSheet.cell(row=row2, column=1).value
                        check = wrSheet.cell(row=row2, column=1)

            # Add highlighter
            yellowHigh = PatternFill(patternType='solid', fgColor=colors.Color(rgb='FFFF00'))
            wrSheet.cell(row=rowY, column=5).fill = yellowHigh
            wrSheet.cell(row=rowY, column=5).value = 'Total Sold'
            wrSheet.cell(row=rowY, column=6).fill = yellowHigh
            wrSheet.cell(row=rowY, column=6).value = str(bottlesSold) # Counted in transcribe loop
            wrSheet.cell(row=rowY, column=7).fill = yellowHigh
            wrSheet.cell(row=rowY, column=7).value = 'Customers'
            wrSheet.cell(row=rowY, column=8).fill = yellowHigh
            wrSheet.cell(row=rowY, column=8).value = str(custEngage) # Counted in transcribe loop

            # Transcribe proper file name
            name1 = weekNum + client.value
            name2 = ' Recap.xlsx'

            nameTitleA = ' Recap '
            nameTitleB = name1 + nameTitleA + dateRange

            name3 = name1 + name2

            wrSheet.cell(row=1, column=1).value = nameTitleB

            wrFile.save(str(name3))

            return 0

        # If next client is different, save file and proceed to new file
        if clNext.value != client.value:

            # Sort data values in client file by date
            sort = False
            #maxRow = wrSheet.max_row + 1
            #maxCol = wrSheet.max_column + 1
            row1 = 3
            row2 = 4
            col1 = 1
            swapCount = 0
            #wrSheet.cell(row=row1, column=1)
            date1 = wrSheet.cell(row=row1, column=1).value
            date2 = wrSheet.cell(row=row2, column=1).value
            check = wrSheet.cell(row=row2, column=1)

            # SKIP SORT IF CLIENT IS ONLY 1 ROW
            if rowCount == 3:
                sort = True
                print('----------------------------------------------')
                print('Skipping sort...')
                print('----------------------------------------------')

            # MANUAL SORT IF CLIENT IS 2 ROWS
            elif rowCount == 4:
                sort = True
                da1 = date1.day
                da2 = date2.day

                # ROW SWAP
                if (da1 > da2):
                    swapCount += 1

                    # SWAP COLUMN DATETIMES
                    col1 = 1
                    temp = wrSheet.cell(row=row1, column=col1).value
                    wrSheet.cell(row=row1, column=col1).value = wrSheet.cell(row=row2, column=col1).value
                    wrSheet.cell(row=row2, column=col1).value = temp

                    # SWAP COLUMN STRINGS
                    for y in range(2, 10):
                        temp = wrSheet.cell(row=row1, column=y).value
                        wrSheet.cell(row=row1, column=y).value = wrSheet.cell(row=row2, column=y).value
                        wrSheet.cell(row=row2, column=y).value = temp
                print('----------------------------------------------')
                print('Manual sort...')
                print('----------------------------------------------')

            print('----------------------------------------------')
            print('Bubble sort engaged...')
            print('----------------------------------------------')
            # Bubble Sort
            while not sort:
                # PASS X
                print('----------------------------------------------')
                print('Pass X for Client ', client.value, 'Row Count: ', rowCount )
                print('----------------------------------------------')
                maxCount = rowCount+1
                for x in range(3, maxCount):
                    print('----------------------------------------------')
                    print('For x in range(3, maxCount): x is ', x, ' | Max Count is', maxCount)
                    print('----------------------------------------------')

                    # ITERATE & LOOK FOR ROWS TO SWAPS
                    if check.value is not None:
                        da1 = date1.day
                        da2 = date2.day

                        print('----------------------------------------------')
                        print('Are Date 1 and 2 correctly ordered...?')
                        print('----------------------------------------------')

                        # ROW SWAP
                        if (da1 > da2):
                            print('----------------------------------------------')
                            print('Swapping rows...')
                            print('----------------------------------------------')
                            swapCount += 1

                            # SWAP COLUMN DATETIMES
                            col1 = 1
                            temp = wrSheet.cell(row=row1, column=col1).value
                            wrSheet.cell(row=row1, column=col1).value = wrSheet.cell(row=row2, column=col1).value
                            wrSheet.cell(row=row2, column=col1).value = temp

                            # SWAP COLUMN STRINGS
                            for y in range(2, 10):
                                temp = wrSheet.cell(row=row1, column=y).value
                                wrSheet.cell(row=row1, column=y).value = wrSheet.cell(row=row2, column=y).value
                                wrSheet.cell(row=row2, column=y).value = temp

                        if (x == maxCount-1 and swapCount == 0):
                            # If no swaps occurred, sort is complete
                            print('----------------------------------------------')
                            print('No swaps occurred, sort is complete!')
                            print('----------------------------------------------')
                            sort = True

                        print('----------------------------------------------')
                        print('Sort: ', sort, 'If its false we hit neither if statements')
                        print('----------------------------------------------')

                        row1 += 1
                        row2 += 1
                        date1 = wrSheet.cell(row=row1, column=1).value
                        date2 = wrSheet.cell(row=row2, column=1).value
                        check = wrSheet.cell(row=row2, column=1)


                    elif swapCount == 0:

                        sort = True

                        swapCount = 0

                    # ELSE, NEW PASS (RESET SWAP SEARCH)
                    else:
                        swapCount = 0
                        print('----------------------------------------------')
                        print('New Sorting Pass, Row 1: ', row1 )
                        print('----------------------------------------------')
                        row1 = 3
                        row2 = 4
                        date1 = wrSheet.cell(row=row1, column=1).value
                        date2 = wrSheet.cell(row=row2, column=1).value
                        check = wrSheet.cell(row=row2, column=1)

            print('----------------------------------------------')
            print('Finished sorting. Adding highlighted totals...')
            print('----------------------------------------------')
            # Add highlighter
            yellowHigh = PatternFill(patternType='solid', fgColor=colors.Color(rgb='FFFF00'))
            wrSheet.cell(row=rowY, column=5).fill = yellowHigh
            wrSheet.cell(row=rowY, column=5).value = 'Total Sold'
            wrSheet.cell(row=rowY, column=6).fill = yellowHigh
            wrSheet.cell(row=rowY, column=6).value = str(bottlesSold) # Counted in transcribe loop
            wrSheet.cell(row=rowY, column=7).fill = yellowHigh
            wrSheet.cell(row=rowY, column=7).value = 'Customers'
            wrSheet.cell(row=rowY, column=8).fill = yellowHigh
            wrSheet.cell(row=rowY, column=8).value = str(custEngage) # Counted in transcribe loop

            # Transcribe proper file name
            name1 = weekNum + client.value
            name2 = ' Recap.xlsx'

            nameTitleA = ' Recap '
            nameTitleB = name1 + nameTitleA + dateRange

            name3 = name1 + name2

            # Add sheet title
            wrSheet.cell(row=1, column=1).value = nameTitleB
            # Add logo
            logoPaste = Image(logo)
            wrSheet.add_image(logoPaste, 'G1')

            wrFile.save(str(name3))

            print('----------------------------------------------')
            print( 'We just processed ', client.value, ' file.')
            print('We process ', clNext.value, ' next.')
            print('----------------------------------------------')

            rowX += 1
            return rowX

        rowX+=1
        client = xlSheet.cell(row=rowX, column=columns)
        clNext = xlSheet.cell(row=rowX+1, column=columns)

"""
def schedule():
    # Load data into openpyxl format
    xlFile = openpyxl.load_workbook(filename)
    xlSheet = xlFile.active

    # Testing
    data1 = xlSheet.cell(row=1, column=1)
    print(data1.value)

    # Create new empty file
    wrFile = openpyxl.Workbook()
    wrSheet = wrFile.active

    # Get column count and first client
    columns = xlSheet.max_column
    client = xlSheet.cell(row=rowX, column=11)
    rowY = rowX + 1
    clNext = xlSheet.cell(row=rowY, column=11)
    rowY = 2

    while client != 0:
        # Transcribe client data [ first row given ONLY ]
        for x in range(1, columns):
            data1 = xlSheet.cell(row=rowX, column=x)

            if data1 == 0:
                # If cell to be copied is empty
                break
            else:
                # Copy data to new sheet
                wrSheet.cell(row=rowY, column=x).value = data1.value
        rowY += 1

        # For testing
        # print(client.value)

        # If next client is empty, done
        if clNext == 0:
            # Transcribe proper file name
            cl = "client"
            clientFileName = f'{cl}{clNum}'
            fileExt = ".xlsx"
            clientFileName += fileExt
            wrFile.save(str(clientFileName))
            return 0

        # If next client is different, save file and return
        if clNext.value != client.value:
            # Transcribe proper file name
            cl = "client"
            clientFileName = f'{cl}{clNum}'
            fileExt = ".xlsx"
            clientFileName += fileExt
            wrFile.save(str(clientFileName))

            # print(clientFileName)

            rowX += 1
            return rowX

        rowX += 1
        client = xlSheet.cell(row=rowX, column=14)
        clNext = xlSheet.cell(row=rowX + 1, column=14)
"""